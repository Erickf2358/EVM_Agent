import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

HEADER_FILL = 'D9E6F2'


def build_template(sheet_title: str, headers: list[str]) -> HttpResponse:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type='solid')
        ws.column_dimensions[get_column_letter(col_idx)].width = max(18, len(header) + 2)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{sheet_title}_template.xlsx"'
    wb.save(response)
    return response


def build_prefilled_template(sheet_title: str, headers: list[str], rows: list[list]) -> HttpResponse:
    """Build an xlsx template with header row plus pre-filled baseline data rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type='solid')
        ws.column_dimensions[get_column_letter(col_idx)].width = max(18, len(header) + 2)

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{sheet_title}_template.xlsx"'
    wb.save(response)
    return response


def read_rows(uploaded_file, expected_headers: list[str]):
    """Read an uploaded xlsx file and yield dict rows keyed by expected_headers.

    Raises ValueError if the header row doesn't match expected_headers.
    """
    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if header_row is None:
        return

    actual_headers = [str(h).strip() if h is not None else '' for h in header_row[:len(expected_headers)]]
    if actual_headers != expected_headers:
        raise ValueError(
            f'Unexpected column headers. Expected {expected_headers}, got {actual_headers}.'
        )

    for row in rows_iter:
        if row is None or all(v is None for v in row):
            continue
        values = list(row[:len(expected_headers)])
        yield dict(zip(expected_headers, values))
